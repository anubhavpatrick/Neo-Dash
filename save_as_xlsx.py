'''
A module to save user generated data in xlsx file
'''

import datetime
import os
from openpyxl import load_workbook, Workbook

def get_file_name_with_path():
    '''
    Returns the file name with path of the xlsx file
    '''
    year = datetime.datetime.now().year
    month = datetime.datetime.now().strftime("%B")
    day_name = datetime.datetime.now().strftime("%A")
    day = datetime.datetime.now().strftime("%d-%m-%y")#"%D").replace("/","-")
    file_path = f"reports/{year}/{month}/"
    file_name = f"{day_name}-{day}.xlsx"
    return (file_path, file_name)

def create_file(filepath, filename):
    '''A method that creates a file if it does not exist'''
    absolute_file_name = filepath + filename
    try:
        f= open(absolute_file_name, 'ab')
        #file already exists, do nothing
        pass
    except FileNotFoundError:
        os.makedirs(os.path.dirname(absolute_file_name), exist_ok=True)
        wb = Workbook()
        wb.create_sheet("Sales_Report")
        wb.save(absolute_file_name)

def save_data(data):
    
    filepath, filename = get_file_name_with_path()
    
    create_file(filepath, filename)

    absolute_file_path = filepath + filename

    wb = load_workbook(absolute_file_path)

    # grab the active worksheet
    ws = wb.active
    
    # iterate over the data and write it out row by row
    for row in data:
        ws.append(row)
    
    wb.save(absolute_file_path)

def unfurl_data(customer_info, products_dict, product_count, total_price, order_no):
    '''A method to create a list of lists to be saved in the xlsx file'''
    data_as_list = list()
    for i in range(1,product_count+1):
        temp = list()
        #order number
        temp.append(order_no)
        temp.append(customer_info["cust_phone"])
        temp.append(customer_info["cust_name"])
        temp.append(customer_info["pet_type"])
        temp.append(customer_info["payment_method"])
        temp.append(products_dict[f"product_category_{i}"])
        temp.append(products_dict[f"product_name_{i}"])
        temp.append(products_dict[f"product_quantity_{i}"])
        temp.append(products_dict[f"original_price_{i}"])
        temp.append(products_dict[f"discount_{i}"])
        temp.append(products_dict[f"final_price_{i}"])
        data_as_list.append(temp)
    return data_as_list


def generate_final_receipt_and_log(customer_info, products_dict, product_count, total_price, order_no):
    #unfurl data into list of lists for writing to xlsx
    unfurled_data = unfurl_data(customer_info, products_dict, product_count, total_price, order_no)
    
    #save data to xlsx
    save_data(unfurled_data)



#save_data({"hello":123})